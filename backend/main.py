"""
资产管理平台 - FastAPI 后端
提供 RESTful API 接口
"""

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
import sqlite3
from datetime import datetime
from pathlib import Path
import uvicorn
import io
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==================== 配置 ====================
DATA_DIR = Path.home() / ".account_book"
DB_FILE = DATA_DIR / "assets.db"

# ==================== 数据模型 ====================
class AssetItem(BaseModel):
    category: str
    name: str
    amount: float

class RecordCreate(BaseModel):
    month: str
    items: List[AssetItem]

class RecordResponse(BaseModel):
    id: int
    month: str
    timestamp: str
    total_assets: float
    total_liabilities: float
    net_assets: float
    items: List[AssetItem]

class TrendData(BaseModel):
    month: str
    total_assets: float
    total_liabilities: float
    net_assets: float
    categories: dict

# ==================== 数据库操作 ====================
def ensure_db():
    """确保数据库目录和表存在"""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            month TEXT NOT NULL,
            timestamp TEXT NOT NULL,
            total_assets REAL DEFAULT 0,
            total_liabilities REAL DEFAULT 0,
            net_assets REAL DEFAULT 0
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            record_id INTEGER NOT NULL,
            category TEXT NOT NULL,
            name TEXT NOT NULL,
            amount REAL NOT NULL,
            FOREIGN KEY (record_id) REFERENCES records(id) ON DELETE CASCADE
        )
    ''')
    
    # 支出记录表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            category TEXT NOT NULL,
            name TEXT NOT NULL,
            amount REAL NOT NULL,
            note TEXT,
            timestamp TEXT NOT NULL
        )
    ''')
    
    # 预算表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS budgets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            month TEXT NOT NULL UNIQUE,
            total_budget REAL DEFAULT 0,
            category_budgets TEXT DEFAULT '{}',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    ''')
    
    conn.commit()
    conn.close()

def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def calculate_totals(items: List[AssetItem]):
    """计算总资产、总负债、净资产"""
    asset_categories = ["现金类", "固定资产", "投资", "重要财产"]
    liability_categories = ["负债类"]
    
    total_assets = sum(item.amount for item in items if item.category in asset_categories)
    total_liabilities = sum(item.amount for item in items if item.category in liability_categories)
    net_assets = total_assets - total_liabilities
    
    return round(total_assets, 2), round(total_liabilities, 2), round(net_assets, 2)

# ==================== FastAPI 应用 ====================
app = FastAPI(title="资产管理平台 API", version="2.0.0")

# CORS 配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def startup():
    ensure_db()

@app.get("/")
def root():
    return {"message": "资产管理平台 API", "version": "2.0.0"}

@app.get("/api/months")
def get_months():
    """获取所有月份列表"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT DISTINCT month FROM records ORDER BY month DESC')
    months = [row['month'] for row in cursor.fetchall()]
    conn.close()
    return {"months": months}

@app.get("/api/records/{month}")
def get_records_by_month(month: str):
    """获取指定月份的记录"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT id, month, timestamp, total_assets, total_liabilities, net_assets
        FROM records WHERE month = ? ORDER BY timestamp DESC
    ''', (month,))
    
    records = []
    for row in cursor.fetchall():
        cursor.execute('''
            SELECT category, name, amount FROM items WHERE record_id = ?
        ''', (row['id'],))
        
        items = [AssetItem(
            category=item['category'],
            name=item['name'],
            amount=item['amount']
        ) for item in cursor.fetchall()]
        
        records.append({
            "id": row['id'],
            "month": row['month'],
            "timestamp": row['timestamp'],
            "total_assets": row['total_assets'],
            "total_liabilities": row['total_liabilities'],
            "net_assets": row['net_assets'],
            "items": items
        })
    
    conn.close()
    return {"records": records}

@app.get("/api/latest")
def get_latest():
    """获取最近一条记录"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT id, month, timestamp FROM records ORDER BY timestamp DESC LIMIT 1
    ''')
    row = cursor.fetchone()
    
    if row:
        cursor.execute('''
            SELECT category, name, amount FROM items WHERE record_id = ?
        ''', (row['id'],))
        
        items = [AssetItem(
            category=item['category'],
            name=item['name'],
            amount=item['amount']
        ) for item in cursor.fetchall()]
        
        conn.close()
        return {"month": row['month'], "items": items}
    
    conn.close()
    return {"month": None, "items": []}

@app.post("/api/records")
def create_record(record: RecordCreate):
    """创建新记录"""
    conn = get_db()
    cursor = conn.cursor()
    
    total_assets, total_liabilities, net_assets = calculate_totals(record.items)
    
    cursor.execute('''
        INSERT INTO records (month, timestamp, total_assets, total_liabilities, net_assets)
        VALUES (?, ?, ?, ?, ?)
    ''', (record.month, datetime.now().isoformat(), total_assets, 
          total_liabilities, net_assets))
    
    record_id = cursor.lastrowid
    
    for item in record.items:
        cursor.execute('''
            INSERT INTO items (record_id, category, name, amount)
            VALUES (?, ?, ?, ?)
        ''', (record_id, item.category, item.name, item.amount))
    
    conn.commit()
    conn.close()
    
    return {
        "success": True,
        "record_id": record_id,
        "totals": {
            "total_assets": total_assets,
            "total_liabilities": total_liabilities,
            "net_assets": net_assets
        }
    }

@app.get("/api/trend")
def get_trend():
    """获取资产趋势数据"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT id, month, timestamp, total_assets, total_liabilities, net_assets
        FROM records ORDER BY month ASC, timestamp DESC
    ''')
    
    monthly_data = {}
    for row in cursor.fetchall():
        month = row['month']
        if month not in monthly_data:
            monthly_data[month] = dict(row)
    
    result = []
    for month, record in monthly_data.items():
        cursor.execute('''
            SELECT category, SUM(amount) as total
            FROM items WHERE record_id = ?
            GROUP BY category
        ''', (record['id'],))
        
        categories = {row['category']: row['total'] for row in cursor.fetchall()}
        
        result.append({
            "month": month,
            "total_assets": record['total_assets'],
            "total_liabilities": record['total_liabilities'],
            "net_assets": record['net_assets'],
            "categories": categories
        })
    
    conn.close()
    return {"data": result}

@app.get("/api/items-trend")
def get_items_trend():
    """获取各资产项趋势数据"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT id, month FROM records ORDER BY month ASC, timestamp DESC
    ''')
    
    monthly_records = {}
    for row in cursor.fetchall():
        month = row['month']
        if month not in monthly_records:
            monthly_records[month] = row['id']
    
    item_data = {}
    for month, record_id in monthly_records.items():
        cursor.execute('''
            SELECT category, name, amount FROM items WHERE record_id = ?
        ''', (record_id,))
        
        for row in cursor.fetchall():
            key = f"{row['category']} - {row['name']}"
            if key not in item_data:
                item_data[key] = {
                    "category": row['category'],
                    "name": row['name'],
                    "data": {}
                }
            item_data[key]['data'][month] = row['amount']
    
    conn.close()
    return {"items": item_data}

@app.delete("/api/records/{record_id}")
def delete_record(record_id: int):
    """删除记录"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM items WHERE record_id = ?', (record_id,))
    cursor.execute('DELETE FROM records WHERE id = ?', (record_id,))
    conn.commit()
    conn.close()
    return {"success": True}

# ==================== 资产健康度评分 ====================
class HealthScoreDetail(BaseModel):
    """健康度评分详情"""
    score: float
    max_score: float
    percentage: float
    level: str
    description: str

class HealthScoreResponse(BaseModel):
    """健康度评分响应"""
    total_score: float
    total_max: float
    grade: str
    grade_description: str
    details: dict
    suggestions: List[str]

def calculate_net_asset_health(net_asset_ratio: float) -> HealthScoreDetail:
    """净资产健康度评分 (25分)"""
    if net_asset_ratio >= 0.8:
        score, level, desc = 25, "优秀", "净资产比例非常健康，负债控制良好"
    elif net_asset_ratio >= 0.6:
        score, level, desc = 20, "良好", "净资产比例健康，负债处于可控范围"
    elif net_asset_ratio >= 0.4:
        score, level, desc = 15, "中等", "净资产比例一般，建议适度降低负债"
    elif net_asset_ratio >= 0.2:
        score, level, desc = 10, "及格", "负债较高，需关注偿债能力"
    else:
        score, level, desc = 5, "警告", "负债过高，存在财务风险"
    
    return HealthScoreDetail(
        score=score,
        max_score=25,
        percentage=score / 25 * 100,
        level=level,
        description=desc
    )

def calculate_liquidity_health(cash_ratio: float) -> HealthScoreDetail:
    """流动性健康度评分 (25分) - 现金类资产占比"""
    if 0.1 <= cash_ratio <= 0.3:
        score, level, desc = 25, "优秀", "流动性储备合理，既能应急又能增值"
    elif 0.05 <= cash_ratio < 0.1 or 0.3 < cash_ratio <= 0.5:
        score, level, desc = 20, "良好", "流动性储备较为合理"
    elif 0.5 < cash_ratio <= 0.7:
        score, level, desc = 15, "中等", "现金占比偏高，可考虑增加投资"
    elif cash_ratio > 0.7:
        score, level, desc = 10, "保守", "现金占比过高，资金利用效率低"
    else:
        score, level, desc = 10, "不足", "流动性储备不足，建议增加现金类资产"
    
    return HealthScoreDetail(
        score=score,
        max_score=25,
        percentage=score / 25 * 100,
        level=level,
        description=desc
    )

def calculate_diversity_health(category_count: int) -> HealthScoreDetail:
    """资产多样性评分 (20分)"""
    if category_count >= 4:
        score, level, desc = 20, "优秀", "资产配置多元化，风险分散良好"
    elif category_count == 3:
        score, level, desc = 15, "良好", "资产配置较为多元"
    elif category_count == 2:
        score, level, desc = 10, "中等", "资产类型较少，建议增加配置多样性"
    else:
        score, level, desc = 5, "不足", "资产类型单一，风险集中度高"
    
    return HealthScoreDetail(
        score=score,
        max_score=20,
        percentage=score / 20 * 100,
        level=level,
        description=desc
    )

def calculate_debt_health(liability_ratio: float) -> HealthScoreDetail:
    """负债控制评分 (15分) - 负债占总资产比例"""
    if liability_ratio < 0.2:
        score, level, desc = 15, "优秀", "负债率很低，财务压力小"
    elif liability_ratio < 0.4:
        score, level, desc = 12, "良好", "负债率适中，财务状况健康"
    elif liability_ratio < 0.6:
        score, level, desc = 8, "中等", "负债率偏高，需关注还款能力"
    elif liability_ratio < 0.8:
        score, level, desc = 5, "警告", "负债率过高，存在财务风险"
    else:
        score, level, desc = 2, "危险", "负债率极高，需立即优化财务结构"
    
    return HealthScoreDetail(
        score=score,
        max_score=15,
        percentage=score / 15 * 100,
        level=level,
        description=desc
    )

def calculate_investment_health(investment_ratio: float) -> HealthScoreDetail:
    """投资配置评分 (15分) - 投资类资产占比"""
    if 0.1 <= investment_ratio <= 0.4:
        score, level, desc = 15, "优秀", "投资比例合理，风险收益平衡"
    elif 0.05 <= investment_ratio < 0.1 or 0.4 < investment_ratio <= 0.6:
        score, level, desc = 10, "良好", "投资比例较为合理"
    elif investment_ratio > 0.6:
        score, level, desc = 5, "激进", "投资比例过高，风险敞口大"
    else:
        score, level, desc = 8, "保守", "投资比例较低，可考虑适度增加"
    
    return HealthScoreDetail(
        score=score,
        max_score=15,
        percentage=score / 15 * 100,
        level=level,
        description=desc
    )

def get_grade(total_score: float) -> tuple:
    """根据总分获取等级"""
    if total_score >= 90:
        return "A+", "财务状况优秀，继续保持！"
    elif total_score >= 80:
        return "A", "财务状况良好，小有改善空间"
    elif total_score >= 70:
        return "B", "财务状况中等，有优化空间"
    elif total_score >= 60:
        return "C", "财务状况一般，建议改进"
    else:
        return "D", "财务状况堪忧，需要重点关注"

def generate_suggestions(details: dict) -> List[str]:
    """根据评分详情生成建议"""
    suggestions = []
    
    # 净资产建议
    if details['net_asset']['score'] < 15:
        suggestions.append("📌 建议逐步降低负债，增加净资产积累")
    
    # 流动性建议
    if details['liquidity']['score'] < 20:
        cash_ratio = details['liquidity']['percentage']
        if cash_ratio < 40:
            suggestions.append("📌 建议增加现金类储备，确保3-6个月的生活应急资金")
        else:
            suggestions.append("📌 现金占比过高，可考虑配置部分理财产品或基金")
    
    # 多样性建议
    if details['diversity']['score'] < 15:
        suggestions.append("📌 建议分散投资，增加资产类型，降低集中风险")
    
    # 负债建议
    if details['debt']['score'] < 10:
        suggestions.append("📌 建议制定还款计划，优先偿还高利率负债")
    
    # 投资建议
    if details['investment']['score'] < 10:
        inv_ratio = details['investment']['percentage']
        if inv_ratio < 50:
            suggestions.append("📌 可适当增加投资配置，让资产增值")
        else:
            suggestions.append("📌 投资比例过高，建议适当降低风险敞口")
    
    if not suggestions:
        suggestions.append("🎉 您的资产配置非常健康，继续保持！")
    
    return suggestions

@app.get("/api/health-score")
def get_health_score():
    """获取资产健康度评分"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 获取最新记录
    cursor.execute('''
        SELECT id, month, timestamp, total_assets, total_liabilities, net_assets
        FROM records ORDER BY timestamp DESC LIMIT 1
    ''')
    row = cursor.fetchone()
    
    if not row:
        conn.close()
        return {
            "has_data": False,
            "message": "暂无资产记录，请先录入资产数据"
        }
    
    record_id = row['id']
    total_assets = row['total_assets']
    total_liabilities = row['total_liabilities']
    net_assets = row['net_assets']
    
    # 获取各类资产金额
    cursor.execute('''
        SELECT category, SUM(amount) as total
        FROM items WHERE record_id = ?
        GROUP BY category
    ''', (record_id,))
    
    category_amounts = {r['category']: r['total'] for r in cursor.fetchall()}
    
    # 获取资产类别数量（不含负债类）
    cursor.execute('''
        SELECT COUNT(DISTINCT category) as count
        FROM items WHERE record_id = ? AND category != '负债类'
    ''', (record_id,))
    category_count = cursor.fetchone()['count']
    
    conn.close()
    
    # 计算各项指标
    asset_categories = ["现金类", "固定资产", "投资", "重要财产"]
    total_asset_value = sum(category_amounts.get(cat, 0) for cat in asset_categories)
    
    cash_amount = category_amounts.get('现金类', 0)
    investment_amount = category_amounts.get('投资', 0)
    
    # 计算比率
    net_asset_ratio = net_assets / total_assets if total_assets > 0 else 0
    cash_ratio = cash_amount / total_asset_value if total_asset_value > 0 else 0
    liability_ratio = total_liabilities / total_assets if total_assets > 0 else 0
    investment_ratio = investment_amount / total_asset_value if total_asset_value > 0 else 0
    
    # 计算各维度评分
    net_asset_health = calculate_net_asset_health(net_asset_ratio)
    liquidity_health = calculate_liquidity_health(cash_ratio)
    diversity_health = calculate_diversity_health(category_count)
    debt_health = calculate_debt_health(liability_ratio)
    investment_health = calculate_investment_health(investment_ratio)
    
    details = {
        "net_asset": net_asset_health.model_dump(),
        "liquidity": liquidity_health.model_dump(),
        "diversity": diversity_health.model_dump(),
        "debt": debt_health.model_dump(),
        "investment": investment_health.model_dump()
    }
    
    # 计算总分
    total_score = (
        net_asset_health.score +
        liquidity_health.score +
        diversity_health.score +
        debt_health.score +
        investment_health.score
    )
    
    grade, grade_description = get_grade(total_score)
    suggestions = generate_suggestions(details)
    
    return {
        "has_data": True,
        "month": row['month'],
        "total_score": round(total_score, 1),
        "total_max": 100,
        "grade": grade,
        "grade_description": grade_description,
        "details": details,
        "suggestions": suggestions,
        "metrics": {
            "net_asset_ratio": round(net_asset_ratio * 100, 1),
            "cash_ratio": round(cash_ratio * 100, 1),
            "liability_ratio": round(liability_ratio * 100, 1),
            "investment_ratio": round(investment_ratio * 100, 1),
            "category_count": category_count
        },
        "summary": {
            "total_assets": total_assets,
            "total_liabilities": total_liabilities,
            "net_assets": net_assets,
            "cash_amount": cash_amount,
            "investment_amount": investment_amount
        }
    }

# ==================== 支出记录功能 ====================
class ExpenseCreate(BaseModel):
    """支出记录创建"""
    date: str  # YYYY-MM-DD
    category: str
    amount: float
    note: Optional[str] = ""

class ExpenseResponse(BaseModel):
    """支出记录响应"""
    id: int
    date: str
    category: str
    name: str
    amount: float
    note: str
    timestamp: str

# 支出分类配置
EXPENSE_CATEGORIES = [
    {"value": "餐饮", "label": "🍜 餐饮", "color": "orange"},
    {"value": "交通", "label": "🚗 交通", "color": "blue"},
    {"value": "服饰", "label": "👕 服饰", "color": "pink"},
    {"value": "购物", "label": "🛒 购物", "color": "purple"},
    {"value": "运动", "label": "🏃 运动", "color": "green"},
    {"value": "娱乐", "label": "🎬 娱乐", "color": "magenta"},
    {"value": "学习", "label": "📚 学习", "color": "cyan"},
    {"value": "医疗", "label": "💊 医疗", "color": "red"},
    {"value": "居住", "label": "🏠 居住", "color": "geekblue"},
    {"value": "社交", "label": "💝 社交", "color": "gold"},
    {"value": "宠物", "label": "🐾 宠物", "color": "lime"},
    {"value": "其他", "label": "💰 其他", "color": "default"},
]

@app.get("/api/expense-categories")
def get_expense_categories():
    """获取支出分类列表"""
    return {"categories": EXPENSE_CATEGORIES}

@app.post("/api/expenses")
def create_expense(expense: ExpenseCreate):
    """创建支出记录"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 获取分类的标签作为显示名称
    category_label = expense.category
    for cat in EXPENSE_CATEGORIES:
        if cat["value"] == expense.category:
            category_label = cat["label"]
            break
    
    cursor.execute('''
        INSERT INTO expenses (date, category, name, amount, note, timestamp)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (expense.date, expense.category, category_label, expense.amount, 
          expense.note or "", datetime.now().isoformat()))
    
    expense_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return {"success": True, "id": expense_id}

@app.get("/api/expenses")
def get_expenses(month: Optional[str] = None, limit: int = 100):
    """获取支出记录列表"""
    conn = get_db()
    cursor = conn.cursor()
    
    if month:
        cursor.execute('''
            SELECT id, date, category, name, amount, note, timestamp
            FROM expenses WHERE strftime('%Y-%m', date) = ?
            ORDER BY date DESC, timestamp DESC
            LIMIT ?
        ''', (month, limit))
    else:
        cursor.execute('''
            SELECT id, date, category, name, amount, note, timestamp
            FROM expenses ORDER BY date DESC, timestamp DESC
            LIMIT ?
        ''', (limit,))
    
    expenses = []
    for row in cursor.fetchall():
        expenses.append({
            "id": row['id'],
            "date": row['date'],
            "category": row['category'],
            "name": row['name'],
            "amount": row['amount'],
            "note": row['note'],
            "timestamp": row['timestamp']
        })
    
    conn.close()
    return {"expenses": expenses}

@app.get("/api/expenses/summary/{month}")
def get_expense_summary(month: str):
    """获取月度支出汇总"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 按分类汇总
    cursor.execute('''
        SELECT category, SUM(amount) as total, COUNT(*) as count
        FROM expenses WHERE strftime('%Y-%m', date) = ?
        GROUP BY category
        ORDER BY total DESC
    ''', (month,))
    
    by_category = []
    total_amount = 0
    for row in cursor.fetchall():
        by_category.append({
            "category": row['category'],
            "total": row['total'],
            "count": row['count']
        })
        total_amount += row['total']
    
    # 按日期汇总
    cursor.execute('''
        SELECT date, SUM(amount) as total, COUNT(*) as count
        FROM expenses WHERE strftime('%Y-%m', date) = ?
        GROUP BY date
        ORDER BY date ASC
    ''', (month,))
    
    by_date = []
    for row in cursor.fetchall():
        by_date.append({
            "date": row['date'],
            "total": row['total'],
            "count": row['count']
        })
    
    conn.close()
    
    return {
        "month": month,
        "total": round(total_amount, 2),
        "by_category": by_category,
        "by_date": by_date,
        "category_count": len(by_category),
        "record_count": sum(item['count'] for item in by_date)
    }

@app.get("/api/expenses/months")
def get_expense_months():
    """获取有支出记录的月份列表"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT DISTINCT strftime('%Y-%m', date) as month
        FROM expenses ORDER BY month DESC
    ''')
    
    months = [row['month'] for row in cursor.fetchall()]
    conn.close()
    return {"months": months}

@app.delete("/api/expenses/{expense_id}")
def delete_expense(expense_id: int):
    """删除支出记录"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM expenses WHERE id = ?', (expense_id,))
    conn.commit()
    affected = cursor.rowcount
    conn.close()
    
    if affected == 0:
        raise HTTPException(status_code=404, detail="支出记录不存在")
    
    return {"success": True}

# ==================== Excel 导出功能 ====================
@app.get("/api/export/assets")
def export_assets():
    """导出资产记录为Excel"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 获取所有记录
    cursor.execute('''
        SELECT id, month, timestamp, total_assets, total_liabilities, net_assets
        FROM records ORDER BY month DESC, timestamp DESC
    ''')
    records = cursor.fetchall()
    
    if not records:
        conn.close()
        raise HTTPException(status_code=404, detail="暂无资产记录")
    
    # 创建Excel
    wb = Workbook()
    
    # === Sheet 1: 资产概览 ===
    ws_summary = wb.active
    ws_summary.title = "资产概览"
    
    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题行
    headers = ["月份", "记录时间", "总资产", "总负债", "净资产"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # 数据行
    for row_idx, record in enumerate(records, 2):
        ws_summary.cell(row=row_idx, column=1, value=record['month']).border = thin_border
        ws_summary.cell(row=row_idx, column=2, value=record['timestamp'][:19]).border = thin_border
        ws_summary.cell(row=row_idx, column=3, value=record['total_assets']).border = thin_border
        ws_summary.cell(row=row_idx, column=4, value=record['total_liabilities']).border = thin_border
        ws_summary.cell(row=row_idx, column=5, value=record['net_assets']).border = thin_border
    
    # 调整列宽
    ws_summary.column_dimensions['A'].width = 12
    ws_summary.column_dimensions['B'].width = 22
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 15
    
    # === Sheet 2: 资产明细 ===
    ws_detail = wb.create_sheet("资产明细")
    
    detail_headers = ["月份", "分类", "名称", "金额"]
    for col, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row_idx = 2
    for record in records:
        cursor.execute('''
            SELECT category, name, amount FROM items WHERE record_id = ?
        ''', (record['id'],))
        
        for item in cursor.fetchall():
            ws_detail.cell(row=row_idx, column=1, value=record['month']).border = thin_border
            ws_detail.cell(row=row_idx, column=2, value=item['category']).border = thin_border
            ws_detail.cell(row=row_idx, column=3, value=item['name']).border = thin_border
            ws_detail.cell(row=row_idx, column=4, value=item['amount']).border = thin_border
            row_idx += 1
    
    ws_detail.column_dimensions['A'].width = 12
    ws_detail.column_dimensions['B'].width = 12
    ws_detail.column_dimensions['C'].width = 20
    ws_detail.column_dimensions['D'].width = 15
    
    conn.close()
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"资产记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = urllib.parse.quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={encoded_filename}"}
    )

@app.get("/api/export/expenses")
def export_expenses(month: Optional[str] = None):
    """导出支出记录为Excel"""
    conn = get_db()
    cursor = conn.cursor()
    
    if month:
        cursor.execute('''
            SELECT id, date, category, name, amount, note, timestamp
            FROM expenses WHERE strftime('%Y-%m', date) = ?
            ORDER BY date DESC, timestamp DESC
        ''', (month,))
    else:
        cursor.execute('''
            SELECT id, date, category, name, amount, note, timestamp
            FROM expenses ORDER BY date DESC, timestamp DESC
        ''')
    
    expenses = cursor.fetchall()
    
    if not expenses:
        conn.close()
        raise HTTPException(status_code=404, detail="暂无支出记录")
    
    # 创建Excel
    wb = Workbook()
    
    # === Sheet 1: 支出明细 ===
    ws_detail = wb.active
    ws_detail.title = "支出明细"
    
    # 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题行
    headers = ["日期", "分类", "名称", "金额", "备注", "记录时间"]
    for col, header in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # 数据行
    for row_idx, expense in enumerate(expenses, 2):
        ws_detail.cell(row=row_idx, column=1, value=expense['date']).border = thin_border
        ws_detail.cell(row=row_idx, column=2, value=expense['category']).border = thin_border
        ws_detail.cell(row=row_idx, column=3, value=expense['name']).border = thin_border
        cell = ws_detail.cell(row=row_idx, column=4, value=expense['amount'])
        cell.border = thin_border
        cell.number_format = '¥#,##0.00'
        ws_detail.cell(row=row_idx, column=5, value=expense['note']).border = thin_border
        ws_detail.cell(row=row_idx, column=6, value=expense['timestamp'][:19]).border = thin_border
    
    # 调整列宽
    ws_detail.column_dimensions['A'].width = 12
    ws_detail.column_dimensions['B'].width = 10
    ws_detail.column_dimensions['C'].width = 15
    ws_detail.column_dimensions['D'].width = 12
    ws_detail.column_dimensions['E'].width = 25
    ws_detail.column_dimensions['F'].width = 20
    
    # === Sheet 2: 分类汇总 ===
    ws_summary = wb.create_sheet("分类汇总")
    
    # 按分类汇总
    if month:
        cursor.execute('''
            SELECT category, SUM(amount) as total, COUNT(*) as count
            FROM expenses WHERE strftime('%Y-%m', date) = ?
            GROUP BY category ORDER BY total DESC
        ''', (month,))
    else:
        cursor.execute('''
            SELECT category, SUM(amount) as total, COUNT(*) as count
            FROM expenses GROUP BY category ORDER BY total DESC
        ''')
    
    summary_headers = ["分类", "总金额", "笔数", "平均每笔"]
    header_fill_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill_blue
        cell.alignment = center_align
        cell.border = thin_border
    
    total_amount = 0
    total_count = 0
    
    for row_idx, row in enumerate(cursor.fetchall(), 2):
        ws_summary.cell(row=row_idx, column=1, value=row['category']).border = thin_border
        cell = ws_summary.cell(row=row_idx, column=2, value=row['total'])
        cell.border = thin_border
        cell.number_format = '¥#,##0.00'
        ws_summary.cell(row=row_idx, column=3, value=row['count']).border = thin_border
        avg = row['total'] / row['count'] if row['count'] > 0 else 0
        cell = ws_summary.cell(row=row_idx, column=4, value=avg)
        cell.border = thin_border
        cell.number_format = '¥#,##0.00'
        
        total_amount += row['total']
        total_count += row['count']
    
    # 添加合计行
    total_row = len(expenses) + 2 if month else len(expenses) + 2
    ws_summary.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws_summary.cell(row=total_row, column=1).border = thin_border
    cell = ws_summary.cell(row=total_row, column=2, value=total_amount)
    cell.font = Font(bold=True)
    cell.border = thin_border
    cell.number_format = '¥#,##0.00'
    ws_summary.cell(row=total_row, column=3, value=total_count).font = Font(bold=True)
    ws_summary.cell(row=total_row, column=3).border = thin_border
    
    ws_summary.column_dimensions['A'].width = 12
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 10
    ws_summary.column_dimensions['D'].width = 15
    
    conn.close()
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    month_str = f"_{month}" if month else ""
    filename = f"支出记录{month_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = urllib.parse.quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={encoded_filename}"}
    )

# ==================== 预算管理功能 ====================
class BudgetCreate(BaseModel):
    """预算创建"""
    month: str
    total_budget: float
    category_budgets: Optional[dict] = {}

@app.get("/api/budgets/{month}")
def get_budget(month: str):
    """获取指定月份的预算"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT id, month, total_budget, category_budgets, created_at, updated_at
        FROM budgets WHERE month = ?
    ''', (month,))
    
    row = cursor.fetchone()
    
    if not row:
        # 获取该月支出
        cursor.execute('''
            SELECT COALESCE(SUM(amount), 0) as total
            FROM expenses WHERE strftime('%Y-%m', date) = ?
        ''', (month,))
        spent = cursor.fetchone()['total']
        conn.close()
        
        return {
            "exists": False,
            "month": month,
            "total_budget": 0,
            "category_budgets": {},
            "spent": spent,
            "remaining": 0
        }
    
    import json
    category_budgets = json.loads(row['category_budgets'] or '{}')
    
    # 获取该月支出
    cursor.execute('''
        SELECT COALESCE(SUM(amount), 0) as total
        FROM expenses WHERE strftime('%Y-%m', date) = ?
    ''', (month,))
    spent = cursor.fetchone()['total']
    conn.close()
    
    return {
        "exists": True,
        "id": row['id'],
        "month": row['month'],
        "total_budget": row['total_budget'],
        "category_budgets": category_budgets,
        "spent": spent,
        "remaining": row['total_budget'] - spent,
        "created_at": row['created_at'],
        "updated_at": row['updated_at']
    }

@app.post("/api/budgets")
def create_or_update_budget(budget: BudgetCreate):
    """创建或更新预算"""
    conn = get_db()
    cursor = conn.cursor()
    
    import json
    now = datetime.now().isoformat()
    category_json = json.dumps(budget.category_budgets, ensure_ascii=False)
    
    # 检查是否已存在
    cursor.execute('SELECT id FROM budgets WHERE month = ?', (budget.month,))
    existing = cursor.fetchone()
    
    if existing:
        cursor.execute('''
            UPDATE budgets 
            SET total_budget = ?, category_budgets = ?, updated_at = ?
            WHERE month = ?
        ''', (budget.total_budget, category_json, now, budget.month))
    else:
        cursor.execute('''
            INSERT INTO budgets (month, total_budget, category_budgets, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?)
        ''', (budget.month, budget.total_budget, category_json, now, now))
    
    conn.commit()
    conn.close()
    
    return {"success": True, "month": budget.month}

@app.get("/api/budgets/{month}/status")
def get_budget_status(month: str):
    """获取预算执行状态"""
    conn = get_db()
    cursor = conn.cursor()
    
    # 获取预算
    cursor.execute('SELECT * FROM budgets WHERE month = ?', (month,))
    budget_row = cursor.fetchone()
    
    if not budget_row:
        conn.close()
        return {"has_budget": False, "month": month}
    
    import json
    category_budgets = json.loads(budget_row['category_budgets'] or '{}')
    total_budget = budget_row['total_budget']
    
    # 获取总支出
    cursor.execute('''
        SELECT COALESCE(SUM(amount), 0) as total
        FROM expenses WHERE strftime('%Y-%m', date) = ?
    ''', (month,))
    total_spent = cursor.fetchone()['total']
    
    # 获取分类支出
    cursor.execute('''
        SELECT category, SUM(amount) as total
        FROM expenses WHERE strftime('%Y-%m', date) = ?
        GROUP BY category
    ''', (month,))
    
    category_spent = {}
    for row in cursor.fetchall():
        category_spent[row['category']] = row['total']
    
    conn.close()
    
    # 计算各项状态
    category_status = []
    for cat, budget_amount in category_budgets.items():
        spent = category_spent.get(cat, 0)
        category_status.append({
            "category": cat,
            "budget": budget_amount,
            "spent": spent,
            "remaining": budget_amount - spent,
            "percent": (spent / budget_amount * 100) if budget_amount > 0 else 0
        })
    
    return {
        "has_budget": True,
        "month": month,
        "total_budget": total_budget,
        "total_spent": total_spent,
        "remaining": total_budget - total_spent,
        "percent": (total_spent / total_budget * 100) if total_budget > 0 else 0,
        "category_status": category_status
    }

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
